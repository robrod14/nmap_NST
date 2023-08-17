import subprocess
import os
import glob
import re
from openpyxl import Workbook

count = 1
categories = ["cat1", "cat2", "cat3"]
wb = Workbook()

def menu():
    """This creates the menu displayed to the user. Here they can select which type of scan they 
	would like to run.

	Returns:
		String: Returns the number (as a string) corresponding to the selection they want to run. 
	"""
    optionsList = ["Ping", "Discovery", "Full", "Speed", "UDP", "Exit"]
    for counter, i in enumerate(optionsList,1):
        print(str(counter) + ".", i)
    print("")
    scan_selection = input("Select the type of scan you want: ")
    return scan_selection

def scan(scan_selection, category_file, output_name):
    """Contains the different types of Nmap scans and all the switches needed to run them.

	Args:
		scan_selection (String): The number of the scan that should be ran.
		category_file (File): This is the text file that holds all the IP addresses to scan 
		based on the category.
		output_name (String): The engagement name that will be used when saving the output 
		of the Nmap scans.
	"""
    if scan_selection == "1":
        result = subprocess.run(["nmap", "-sn", "-oA", 
				 output_name + "_PING_" + category_file, "-vvv", 
				 "--open", "-iL", "../" + category_file])
    elif scan_selection == "2":
       result = subprocess.run(["nmap", "-Pn", "-sS", "-p",
				 "21-23,25,53,111,137,139,445,80,443,8443,8080", "--min-hostgroup", 
				 "255", "--min-rtt-timeout", "0ms", "--max-rtt-timeout", "100ms", 
				 "--max-retries", "1", "--max-scan_selection-delay", "0", "--min-rate", 
				 "620", "-oA", output_name + "_DISCOVERY_" + category_file, "-vvv", 
				 "--open", "-iL", "../" + category_file])
    elif scan_selection == "3":
        subprocess.run(["nmap", "-Pn", "-sS", "-p-", "-oA", 
			output_name + "_FULL_" + category_file, "-vvv", "--open", "-iL", 
			"../" + category_file])
    elif scan_selection == "4":
        result = subprocess.run(["nmap", "-Pn", "-n", "-sS", "-p-", 
				 "--min-hostgroup", "255", "--min-rtt-timeout", "0ms", 
				 "--max-rtt-timeout", "100ms", "--max-retries", "1", 
				 "--max-scan_selection-delay", "0", "--min-rate", "620", "-oA", 
				 output_name + "_SPEED_" + category_file, "-vvv", "--open", "-iL", 
				 "../" + category_file])
    else:
        print("you didn't select a valid option")
        scan_selection = menu()

def getFiles():
    """Gets all of the .txt files in the directory.

    Returns:
        List: All of the file names that are in the directory and end with.txt will be returned
	"""
    file_list = os.listdir()
    text_files = [file for file in file_list if file.endswith(".txt")]
    return text_files

def create_report():
	"""Creates the Excel report to send over. Stores the IP Address, Port, and Service. Also puts each category on their own tab.
	"""
	for category in categories:
		files = glob.glob(f"*{category}*.gnmap")
		if files:
			sheet = wb.create_sheet(title=category)
			sheet.append(["IP Address", "Port", "Service"])
			for file in files:
				with open(file, "r") as f:
					content = f.readlines()
				for line in content:
					if "Host:" in line and "Ports:" in line:
						for l in line.split("Ports: ")[1].split(", "):
							if "/open/" in l:
								if "Ignored State" in l:
									pair = (line.split("Host: ")[1].split(" (")[0],re.split("/open/[^\s]*?//", l)[0],re.split("/open/[^\s]*?//", l)[1].split("\tIgnored State:")[0]+"\n")
								else:
									pair = (line.split("Host: ")[1].split(" (")[0],re.split("/open/[^\s]*?//", l)[0],re.split("/open/[^\s]*?//", l)[1])
								sheet.append(pair)
	if "Sheet" in wb.sheetnames:
		wb.remove(wb["Sheet"])
		excel_filename = "report.xlsx"
		wb.save(excel_filename)

	

scan_selection = menu()
output_name = input("What is the name of this engagement? ")
while count < 4:
	ranges = input(f"Insert a comma separated CIDR notation range for all Category {count} IP Addresses: ")
	ranges = ranges.replace(" ", "")
	if ranges:
		with open("cat" + str(count) + ".txt", "w") as f:
			for CIDR_range in ranges.split(","):
				f.write(CIDR_range + "\n")
	count += 1
		

while scan_selection != "6":
	text_files = getFiles()
	os.mkdir("nmap")
	os.chdir("nmap")
	for category_file in text_files:
		scan(scan_selection, category_file, output_name)
	create_report()
	scan_selection = menu()
print("Exiting...")
